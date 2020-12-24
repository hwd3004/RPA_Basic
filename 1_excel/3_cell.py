from random import randint
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws.title = "시트"

# 셀에 데이터 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# A1 셀 출력, 셀의 객체 정보가 뜸
print(ws["A1"])

# A1 셀의 값 출력
print(ws["A1"].value)

# 값이 없을 땐 None 출력
print(ws["A10"].value)

############################################################

# row = 1, 2, 3...
# column = A(1), B(2), C(3)...

# ws["A1"].value
print(ws.cell(column=1, row=1).value)

# ws["B1"].value
print(ws.cell(column=2, row=1).value)

# 이 코드와 같음 ws["C1"] = 10
c = ws.cell(column=3, row=1, value=10)

print(c.value)

##########################################

index = 1

# 10개 row
for x in range(1, 11):
    # 10게 column
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
