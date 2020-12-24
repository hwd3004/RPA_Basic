from openpyxl import load_workbook
import datetime

# 수식 - 데이터 전용

# wb = load_workbook("sample_formula.xlsx")
wb = load_workbook("sample_formula.xlsx", data_only=True)

ws = wb.active

# ws.values로 하면 각 셀에 대해서, 셀 객체가 아닌 밸류를 바로 가져옴
# wb = load_workbook("sample_formula.xlsx", data_only=True)
# 코드 수정
#
# 코드 돌리면 None이 뜨는 이유는,
# 12_formula.py에서 코드 돌린 이후 수식이 그대로 있는 상태에서,
# 엑셀 파일을 열 때 수식이 실행되어서 데이터가 나타난다
# 그런데 지금 13_formula_dataonly.py에서 코드 돌리면
# 엑셀 파일을 열어서 수식이 실행된게 아니기 때문에, None이 뜬다
#
# 이런 현상을 evaluate 라고 한다
# evaluate - 계산 되지 않은 상태의 데이터는 None 처리
#
# 엑셀 파일을 실행시켜서 저장하고 종료 후에 코드를 다시 도릴면 데이터가 출력된다

for row in ws.values:
    for cell in row:
        print(cell)

# 결과

# 2020-12-24 00:39:35.729568
# =sum(1, 2, 3)
# =average(1, 2, 3)
# 10
# 20
# =sum(A4:A5)

# 데이터를 가져오는게 아니라 수식이 뜸

#####################################
