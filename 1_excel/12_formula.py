from openpyxl import Workbook
import datetime

wb = Workbook()

ws = wb.active

# formula - 수식

# 오늘 날짜 정보
ws["A1"] = datetime.datetime.today()

# 1 + 2 + 3 합계
ws["A2"] = "=sum(1, 2, 3)"

# 평균
ws["A3"] = "=average(1, 2, 3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=sum(A4:A5)"

wb.save("sample_formula.xlsx")
