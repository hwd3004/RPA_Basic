from openpyxl import load_workbook

wb = load_workbook("quiz.xlsx", data_only=True)
ws = wb.active

student_data = []
student_obj = []
grade = []

#############################################################################################################


def get_student_data():
    for i in ws.values:
        student_data.append(i)

#############################################################################################################


def set_student_point():
    for i in range(1, len(student_data)):
        # print("출석", student_data[i][1], "| 총점", student_data[i][7])

        obj = {
            "출석": student_data[i][1],
            "총점": student_data[i][7]
        }

        student_obj.append(obj)


##############################################################################################################

def check_student_obj():
    for i in student_obj:
        print(i)

##############################################################################################################


def set_grade():
    for i in student_obj:
        # print(i["출석"])
        if i["출석"] >= 5:
            # print(True)
            if i["총점"] >= 90:
                grade.append("A")
            elif i["총점"] >= 80:
                grade.append("B")
            elif i["총점"] >= 70:
                grade.append("C")
            else:
                grade.append("D")
        elif i["출석"] < 5:
            # print(False)
            grade.append("F")

    print(grade)


##############################################################################################################

def input_grade():
    row_num = 2

    for i in grade:
        ws.cell(column=9, row=row_num, value=i)
        row_num = row_num + 1


##############################################################################################################


def init():
    get_student_data()
    set_student_point()
    check_student_obj()
    set_grade()
    input_grade()
    wb.save("quiz_continue.xlsx")


init()
