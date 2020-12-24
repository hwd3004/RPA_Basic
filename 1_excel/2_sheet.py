from openpyxl import Workbook

wb = Workbook()

# 새로운 시트를 기본 이름으로 생성
ws = wb.create_sheet()

# 시트 이름 변경
ws.title = "시트2"

# RGB 형태로 값을 넣어주면, 색상 변경
ws.sheet_properties.tabColor = "ff66ff"

##########################################

# 해당 이름으로 시트 생성
# 여기까지 하면 sheet, 시트2, 시트3
# 이 3개가 생성됨
ws1 = wb.create_sheet("시트3")

# 2번째 인덱스에 시트 생성
ws2 = wb.create_sheet("시트4", 2)

# 딕셔너리 형태로 시트에 접근
new_ws = wb["시트4"]

# 모든 시트 이름 확인
print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "복사된 시트"

wb.save("sample.xlsx")
