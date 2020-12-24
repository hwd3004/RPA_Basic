from openpyxl.utils.cell import coordinate_from_string
from random import randint
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])

for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어 column만 가져오기
col_B = ws["B"]

for cell in col_B:
    print(cell.value)

# 영어, 수학 둘 다 가져오기
col_range = ws["B:C"]

for cols in col_range:
    for cell in cols:
        print(cell.value)

# 1번째 row만 가지고 오기
row_title = ws[1]

for cell in row_title:
    print(cell.value)

# 1번쨰 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
row_range = ws[2:6]

for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# 2번쨰 줄부터 마지막 줄까지
row_range = ws[2:ws.max_row]

for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")

        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

# 전체 rows
# print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value)

print()

# 전체 columns
# print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

# 전체 row
for row in ws.iter_rows():
    print(row[1].value)

print()
print()
print()

# 전체 column
# for column in ws.iter_cols():
#     print(column[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번쨰 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

# 세로로 한줄씩 가져오고 싶을땐 iter_rows, 열은 iter_columns

wb.save("sample.xlsx")
